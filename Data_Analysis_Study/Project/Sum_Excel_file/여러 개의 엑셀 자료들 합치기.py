import glob
from openpyxl import load_workbook
from openpyxl import Workbook

# 동일한 내용의 엑셀 파일들을 모두 가져오기
language= glob.glob(r'excel\*.xlsx')

print(language)

language_list=[]
language_IDE=[]
language_use=[]

for lang in language:
    wb=load_workbook(lang, data_only=True)
    ws=wb.active
    language_list.append(str(ws["B1"].value))
    language_IDE.append(ws["B2"].value)
    language_use.append(ws["B3"].value)

# permission denied: 엑셀 파일을 열고 있을 때는 실행 불가

print(language_list)
print(language_IDE)
print(language_use)

try:
    # data_only=True시 엑셀 수식값이 아닌 값을 가져온다
    wb=load_workbook(r'excel\결과.xlsx',data_only=True)
    ws=wb.active

except:
    wb=Workbook()
    ws=wb.active

for i in range(len(language_list)):
    ws.cell(row=i + 1, column=1).value = language_list[i]
    ws.cell(row=i + 1, column=2).value = language_IDE[i]
    ws.cell(row=i + 1, column=3).value = language_use[i]

wb.save(r'excel\결과.xlsx')